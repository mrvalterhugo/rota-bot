# openpyxl needs to be 2.6.2 version
#install using:
# pip install --user -U openpyxl==2.6.2
# Also needs requests
import openpyxl
import requests
import urllib3
import sys
import datetime
from _datetime import date
from _datetime import datetime

#Excel file must be in the same folder.
file = "Rota.xlsx"
rota = openpyxl.load_workbook(file)
sheet = rota['2019-20']

#Get today's date
now = datetime.now().time().hour
today = date.today().day
month = date.today().month
d = datetime.today().date()
t = d.strftime('%d-%m-%Y')

#Tech names by position
Techs = ["H", "I", "J", "K", "L", "O", "P", "Q", "R", "S"]

#Open a text file to receive content
file = open('rota.txt', 'w', encoding="utf-8")

#Here I am using fake names and logins for privacy
login = {'Peter': '@peter1', 'Vitor': '@vitor2', 'Jam': '@jamhk',
         'Jey': '@jeyf', 'Mat': '@mattr', 'Ali': '@cali',
         'Fran': '@francost', 'Aaron': '@aaron', 'Regis': '@regi',
         'Moh': '@mohgn', 'Helen': '@hele'
         }

for rowNum in range(1300, sheet.max_row):    # skip the first 1200 rows
    row = sheet.cell(row=rowNum, column=4)
    produceName = sheet.cell(row=rowNum, column=4).value
    if produceName.strftime('%d-%m-%Y') == t :
        text = "/md \U0001F41D**Tech**\U0001F41D\nPOC's for " + t + ":" + "\n\n"
        file.write(text)
        for T in Techs:
            Position = T + str(2)
            if sheet[T + str(rowNum)].value != None:
                name_value = sheet[Position].value
                Name = name_value + " " + login[name_value]
                Shift = sheet[T + str(rowNum)].value
                if "WFH" in Shift:
                    file.write(Name + "\n")
                    file.write(Shift + " - \U0001F3E0**Working From Home:**\U0001F3E0" + "\n\n")
                elif "AL" in Shift:
                    file.write(Name + "\n")
                    file.write(" \U0001F3D6 **On Annual Leave**" + "\n\n")
                elif "L -" in Shift:
                    file.write(Name + "\n")
                    file.write(Shift + " - \U0001F303** Lates - On-site**\U0001F303" + "\n\n")
                elif "E -" in Shift:
                    file.write(Name + "\n")
                    file.write(Shift + " - \U0001F305** Early - On-site**\U0001F305" + "\n\n")
                elif "D -" in Shift:
                    file.write(Name + "\n")
                    file.write(Shift + " - \U0001F305** Early - On-site**\U0001F305" + "\n\n")
                else:
                    file.write(Name + "\n")
                    file.write(Shift + " - **On-site:**\U0001F3ED" + "\n\n")
        break

file.close()
read_file = open('rota.txt', 'r', encoding="utf-8")

urllib3.disable_warnings(urllib3.exceptions.InsecureRequestWarning)
#Replace this URL for the one you get from Amazon Chime:
whurl = "https://hooks.chime.aws/incomingwebhooks/ec8bd1b5-5541-47f8-9dd8-e1410de27df7?token=EXAMPLETOKENMUY0Vi1aVTJCTnhsN1FrZTF6R3loLWY4U2QteV9SenA4dkk4"
headers = {"Content-Type": "application/json"}
session = requests.Session()
session.post(whurl, headers=headers, json={"Content": read_file.read()})
read_file.close()